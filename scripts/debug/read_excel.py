import openpyxl
import json
import os

def excel_to_json(file_path):
    workbook = openpyxl.load_workbook(file_path)
    data = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = []
        for row in sheet.iter_rows(values_only=True):
            sheet_data.append(row)
        data[sheet_name] = sheet_data
    return data

file_path = "C:/Users/Az09.Team/Git-projects/demo_phong_tro_test/File Test.xlsx"
if os.path.exists(file_path):
    data = excel_to_json(file_path)
    print(json.dumps(data, indent=2, ensure_ascii=False))
else:
    print(f"File not found: {file_path}")
