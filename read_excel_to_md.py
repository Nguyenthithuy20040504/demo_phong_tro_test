import openpyxl
import os

def excel_to_md(file_path):
    workbook = openpyxl.load_workbook(file_path)
    md_content = ""
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        md_content += f"## Sheet: {sheet_name}\n\n"
        
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
            
        # Get headers (first row)
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        md_content += "| " + " | ".join(headers) + " |\n"
        md_content += "| " + " | ".join(["---"] * len(headers)) + " |\n"
        
        # Get data (remaining rows)
        for row in rows[1:]:
            # Check if row is mostly empty
            if all(cell is None for cell in row):
                continue
            cells = [str(cell) if cell is not None else "" for cell in row]
            md_content += "| " + " | ".join(cells) + " |\n"
        md_content += "\n"
    return md_content

file_path = "C:/Users/Az09.Team/Git-projects/demo_phong_tro_test/File Test.xlsx"
output_path = "C:/Users/Az09.Team/Git-projects/demo_phong_tro_test/test_requirements.md"

if os.path.exists(file_path):
    md_content = excel_to_md(file_path)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"Generated {output_path}")
else:
    print(f"File not found: {file_path}")
